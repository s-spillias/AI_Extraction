#####
your_name_here = "Rowan" # or 'Rowan', 'Fabio', 'Matt'
####

bespoke_columns = {"Scott":[5,9,11],"Rowan":[12],'Fabio':[4,10],'Matt':[2,3,8]}
# Context: Pull out the relevant bit of the paper. Did they pull out something relevant to the question
# Response: Given that context, is the response relevant/justified. Just rate the AI
# How similar are the two answers that have been extracted?
# we assume that there is 

import pandas as pd
from tkinter import *
from tkinter import ttk
import openpyxl
import shutil
import os
import random

project_directory = "CBFM/"
input_directory = project_directory + 'Extraction_Output/'
output_path = project_directory + 'Assessement_Output/output_assessment.xlsx'
# Load the Excel spreadsheet
xls_human = openpyxl.load_workbook(input_directory +'Extraction_Human.xlsx')

# Load the Excel file using openpyxl
xls_ai = openpyxl.load_workbook(input_directory +'Extraction_ai.xlsx')




global resume
# Check if 'CBFM/output_assessment.xlsx' exists, if not, copy 'CBFM/Extraction_ai.xlsx' to 'CBFM/output_assessment.xlsx'
if not os.path.exists(output_path):
    print("copying source file")
    shutil.copyfile(input_directory +'Extraction_ai.xlsx', output_path)
    resume = False
else:
    print(output_path + " exists")
    resume = True


# Extract the questions from the 'CBFM/criteria.csv' file
questions = ['Response Relevance',"Context Relevance","Content Comparison"]

# Create a tkinter window
root = Tk()
root.resizable(True, True)

# Define a global variable for font size
font_size = 12

# Function to increase font size
def increase_font_size():
    global font_size
    font_size += 2
    for widget in root.winfo_children():
        if isinstance(widget, (ttk.Label, Label)):
            widget.config(font=("Helvetica", font_size))

# Function to decrease font size
def decrease_font_size():
    global font_size
    font_size -= 2
    for widget in root.winfo_children():
        if isinstance(widget, (ttk.Label, Label)):
            widget.config(font=("Helvetica", font_size))

# Add a new Label at the top of the window
rating_label = None
if rating_label is None:
    rating_label = ttk.Label(root, text="1 - Human response is much better \n2 - Responses are similar \n3 - AI response is much better",  font=("Helvetica", 12))
    rating_label.pack()

# Add a new Label at the top of the window


def sanitize_text(text):
    # Trim leading and trailing whitespace
    sanitized = text.strip()
    
    # Replace single quotes with double quotes
    sanitized = sanitized.replace("'", '"')


    # Replace newline characters with a space
    #sanitized = sanitized.replace('\n', ' ')

    # Remove extra internal whitespace by splitting and rejoining
    sanitized = ' '.join(sanitized.split())
    sanitized = sanitized.replace("Context:", '\n\nContext:\n')
    #print(sanitized)
    return sanitized

global suppress
suppress = False
# Function to yield one set of cells at a time, along with the current sheet name and cell position
def cell_generator():
    newline = "\n\n"
    sheet_name = "ALL"
    human_sheet = xls_human[sheet_name]
    max_column_human = human_sheet.max_column
    column_names = [human_sheet.cell(row=1, column=col).value for col in range(1, max_column_human + 1)]
    for col in bespoke_columns[your_name_here]:
        for row in range(2, human_sheet.max_row + 1): 
            ai_sheetnames = xls_ai.sheetnames
            random.shuffle(ai_sheetnames)
            for sheet_name_ai in ai_sheetnames:
                sheet_ai = xls_ai[sheet_name_ai]

                col_name = column_names[col-1] + newline
            
                cell_value_ai = sheet_ai.cell(row=row, column=col).value
                ai_text = sanitize_text(cell_value_ai)
                #print(ai_text)
                human_text = human_sheet.cell(row=row, column=col).value
                human_text = sanitize_text(human_text)
                if ai_text is not None:
                    #print(cell_value_ai[:10])
                    #print(human_text[:10])
                    wb = openpyxl.load_workbook(output_path)
                    sheet = wb[sheet_name_ai]
                    cell = sheet.cell(row=row, column=col)
                    #print(cell)
                    global resume
                    global suppress
                    if len(cell.value.split(";")) == 3 and len(cell.value) == 5:
                        print("Catching up")
                        resume = True
                    else:
                        suppress = True
                        resume = False
                    #print("New cell.")
                    yield f'Response: {human_text} && {newline}    {ai_text}', sheet_name_ai, (row, col),col_name
# Create a generator object
cell_gen = cell_generator()
#print(next(cell_gen))
# Create an area for displaying questions
question_label = Label(root, text='')
question_label.pack()


# Create a list of StringVar objects to store the selected radio button values
selected_values = [StringVar() for _ in range(len(questions))]

first_set_labels = ["1", "2", "3",""]
labels = ["1", "2", "3",""]
# Create a Frame, question Label, and five radio buttons for each question
for i in range(len(questions)):
    frame = ttk.Frame(root)
    frame.pack(padx=5, pady=1)  # Reduce padding around the frame
    question_label = ttk.Label(frame, text=questions[i], font=("Helvetica", 14))
    question_label.grid(row=0, column=0)
    for j in range(1, 4):
        if i == 0:
            radio_button = ttk.Radiobutton(frame, text=first_set_labels[j-1], variable=selected_values[i], value=str(j))
        else:
            radio_button = ttk.Radiobutton(frame, text=labels[j-1], variable=selected_values[i], value=str(j))
        radio_button.grid(row=0, column=j, sticky='E')  # Make the radio buttons right justified
    root.update()

# Create a submit button
def submit():
    global sheet_name_ai, cell_position  # Declare these as global variables
    print("Submit function called")
    while resume:
        print("in submit")
        cell_values, sheet_name_ai, cell_position,col_name = next(cell_gen)
    # Check if all radio buttons have a selected value
    if all(value.get() != '' for value in selected_values):
        # Move to the next cell
        try:
            # Update the specific cell in the DataFrame
            wb = openpyxl.load_workbook(output_path)
            sheet = wb[sheet_name_ai]
            # Unpack the tuple
            row_number , column_number = cell_position
            cell = sheet.cell(row=row_number, column = column_number)
            cell.value = ';'.join(value.get() for value in selected_values)
            # Save the DataFrame back to the Excel file
            wb.save(output_path)
            print("Data written to " + output_path)
            # Reset the selected radio button values
            for value in selected_values:        
                value.set('')
            # Update the display with the new cell data
            cell_values, sheet_name_ai, cell_position, col_name = next(cell_gen)
            update_cells(cell_values, sheet_name_ai, cell_position,col_name)
            # Call the function to update the question   
            #update_question()
            # Force the tkinter window to update 
            root.update()
        except StopIteration:
            print("No more items in generator.")
    else:
        print("Please select a value for each question before submitting.")

submit_button = ttk.Button(root, text="Submit")
submit_button["command"] = submit
# Add Zoom In and Zoom Out buttons to a new Frame
zoom_frame = ttk.Frame(root)
zoom_frame.pack(side='top', anchor='ne')
zoom_in_button = ttk.Button(zoom_frame, text="Zoom In", command=increase_font_size)
zoom_in_button.pack(side='left')
zoom_out_button = ttk.Button(zoom_frame, text="Zoom Out", command=decrease_font_size)
zoom_out_button.pack(side='left')

submit_button.pack()

# Function to yield each question one at a time
def question_generator():
    for question in questions:
        yield question

# Create a generator object
gen = question_generator()

def update_question():
    try:
        # Get the next question
        next_question = next(gen)
    except StopIteration:
        # No more questions
        return
    # Update the question_label with the next question
    question_label.config(text=next_question)

# Start the update process
#update_question()

# Define column_names_label outside the function
column_names_label = None

# Function to update the tkinter window with the current set of cells
def update_cells(cell_values, sheet_name, cell_position,col_name):
    global column_names_label  # Declare column_names_label as global so we can modify it

    # Clear the previous column_names_label
    if column_names_label is not None:
        column_names_label.pack_forget()

    # try:
    #     # Get the next set of cells
    #     #cell_values, sheet_name, cell_position,col_name = next(cell_gen)
    # except StopIteration:
    #     # No more sets of cells
    #     return 

    # Create a new column_names_label and align it to the left
    column_names_label = ttk.Label(root, text="\n\n" + f'{col_name.split(".")[0]}' + "",font=("Helvetica", 16))
    column_names_label.pack(anchor='w')

    # Clear the tkinter window of Label widgets
    for widget in root.winfo_children():
        if isinstance(widget, Label) and widget is not column_names_label:
            widget.pack_forget()

    # Split the cell values by '***'
    cell_values_parts = cell_values.split('&&')
    # Display the first part of the cell values in green with a larger font size, left alignment, and a fixed width
    Label(root, text=cell_values_parts[0], fg='green', wraplength=1200, font=("Helvetica", 12), anchor='w', justify=LEFT, width=100).pack(fill='x', anchor='w')
    # Display the second part of the cell values in red with a larger font size, left alignment, and a fixed width
    Label(root, text=cell_values_parts[1], fg='red', wraplength=1200, font=("Helvetica", 12), anchor='w', justify=LEFT, width=100).pack(fill='x', anchor='w')

    # Call the function to update the question
    #update_question()

while resume:
    #print("in skipping initialization")
    cell_values, sheet_name_ai, cell_position,col_name = next(cell_gen)

# Start the update process
if not resume and not suppress:
    #print("in original start-up")
    cell_values, sheet_name_ai, cell_position,col_name = next(cell_gen)
    suppress = False

update_cells(cell_values, sheet_name_ai, cell_position,col_name)

# Run the tkinter window
root.mainloop()
