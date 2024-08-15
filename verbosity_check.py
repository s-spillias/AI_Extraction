
import pandas as pd
import openpyxl
import shutil
import os
import random
from auxiliary import text_match, full_text

project_directory = "CBFM/"
pdf_location = 'Extraction_Input/pdfs/Human_Kept/'
pdf_location = project_directory + pdf_location
input_directory = project_directory + 'Extraction_Output/'
output_path = project_directory + 'Assessment_Output/output_wordcount.xlsx'
# Load the Excel spreadsheet
xls_human = openpyxl.load_workbook(input_directory +'Extraction_human.xlsx')

# Load the Excel file using openpyxl
xls_ai = openpyxl.load_workbook(input_directory +'Extraction_ai.xlsx')


def word_count(string):
    response,context = string.split("Context:")
    response_count = len(response.split(" "))
    context_count = len(context.split(" "))
    return response_count,context_count

def to_check(context,paper):
    global project_directory
    # Create a DataFrame with the strings
    output_path = project_directory + 'Assessment_Output/output_to_check.xlsx'
    data = {'Paper': [paper], 'Context': [context]}
    df = pd.DataFrame(data)

    # Specify the output path
    sheet_name = "to_check"  # Change this to the name of your sheet

    # Check if the Excel file exists
    if not os.path.exists(output_path):
        # If the file doesn't exist, create a new Excel file with the DataFrame
        df.to_excel(output_path, index=False, sheet_name=sheet_name)
        print("New Excel file created.")
    else:
        # If the file already exists, load the existing data
        existing_data = pd.read_excel(output_path, sheet_name=sheet_name)
        
        # Concatenate the existing DataFrame with the new DataFrame
        updated_data = pd.concat([existing_data, df], ignore_index=True)
        
        # Write the updated DataFrame back to the Excel file
        updated_data.to_excel(output_path, index=False, sheet_name=sheet_name)
        print("Excel sheet updated successfully.")


sheet_name = xls_human.sheetnames[0]
human_sheet = xls_human[sheet_name]
max_column_human = human_sheet.max_column
column_names = [human_sheet.cell(row=1, column=col).value for col in range(1, max_column_human + 1)]
for col in range(2, max_column_human):
    for row in range(2, human_sheet.max_row + 1): 

        for sheet_name_ai in xls_ai.sheetnames:
            response,context = None,None
            sheet_ai = xls_ai[sheet_name_ai]   
            paper = pdf_location + sheet_ai.cell(row=row,column=1).value + '.pdf'
            contents = full_text(paper)
            ai_text = sheet_ai.cell(row=row, column=col).value
            # ai_text = sanitize_text(cell_value_ai)
            human_text = human_sheet.cell(row=row, column=col).value
            # human_text = sanitize_text(human_text)
            ai_word_count = word_count(ai_text)
            human_word_count = word_count(human_text)
            response,context = ai_text.split("Context:")
            matches = text_match(context,contents)
            if not any(matches):
                if "NO CONTEXT" in context or "NO DATA" in context:
                    print("skip")
                else:
                    to_check(context,paper)
            ratio_word_count = [a / b for a, b in zip(ai_word_count,human_word_count)]
            selected_values = [ai_word_count,human_word_count,ratio_word_count]
            flattened_list = [str(item) for sublist in selected_values for item in sublist]

            # Concatenate into a single string
            concatenated_string = ';'.join(flattened_list)

            
            if os.path.exists(output_path):
                wb = openpyxl.load_workbook(output_path)
                # Workbook exists, continue with your operations
            else:
                wb = openpyxl.Workbook()
                wb.save(output_path)
            if sheet_name_ai in wb.sheetnames:
                sheet = wb[sheet_name_ai]
                # Worksheet exists, continue with your operations
            else:
                # Worksheet does not exist, create it
                sheet = wb.create_sheet(title=sheet_name_ai)
                # Optionally, perform any initialization or setup for the new sheet
                # e.g., setting column widths, adding headers, etc.
                
                # Save the workbook after creating the new sheet
                wb.save(output_path)
            # Unpack the tuple
            row_number , column_number = (row, col)
            cell = sheet.cell(row=row_number, column = column_number)
            cell.value = concatenated_string
            # Save the DataFrame back to the Excel file
            wb.save(output_path)