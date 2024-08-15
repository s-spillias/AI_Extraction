
import re
import fitz

import numpy as np
import os
import openpyxl
import pandas as pd
#import pytesseract
#from pdf2image import convert_from_path
#rom fuzzywuzzy import fuzz,process
import openai
import time
from langchain.chat_models import AzureChatOpenAI, ChatOpenAI
from langchain.embeddings import OpenAIEmbeddings
from langchain.document_loaders import PyMuPDFLoader
from PyPDF2 import PdfReader
from langchain.text_splitter import RecursiveCharacterTextSplitter
from dotenv import load_dotenv
import random
import string
from collections import Counter
from fuzzywuzzy import fuzz
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.metrics.pairwise import cosine_similarity

load_dotenv()

openai.api_key = os.environ.get("OPENAI_API_KEY")

def configure_azure_support():
    azure_openai_key = os.getenv("AZURE_OPENAI_KEY")
    azure_openai_endpoint = os.getenv("AZURE_OPENAI_ENDPOINT")
    azure_openai_version = os.getenv("AZURE_OPENAI_VERSION")
    if None in [
        azure_openai_key,
        azure_openai_endpoint,
        azure_openai_version,
    ]:
        print(f"Error: missing Azure OpenAI environment variables. Please see README section on Azure.")
        return

    openai.api_type = "azure"
    openai.api_key = azure_openai_key
    openai.api_base = azure_openai_endpoint
    openai.api_version = azure_openai_version
    
def add_rand_string(question):
    rand_string = ''.join(random.choice(string.ascii_letters + string.digits) for _ in range(20))
    new_question = f"Ignore this string: {rand_string}" + "\n\n" + question             
    return new_question

def get_llm():
    
    load_dotenv()

    llm = AzureChatOpenAI(
        azure_deployment=os.environ.get("AZURE_OPENAI_DEPLOYMENT"),
        openai_api_base=os.environ.get("AZURE_OPENAI_ENDPOINT"),
        openai_api_type="azure",
        openai_api_version=os.environ.get("AZURE_OPENAI_VERSION"),
        openai_api_key=os.environ.get("AZURE_OPENAI_KEY"),
        model_name="gpt-4",
        temperature= os.environ.get("TEMPERATURE")
    )
    # llm = AzureOpenAI(deployment_name='firstcontact-gpt4',
    #                   model_name = "gpt-4",
    #                   openai_api_key=os.environ.get("AZURE_OPENAI_KEY"),
    #                   openai_api_version=os.environ.get("AZURE_OPENAI_VERSION"),
    #                   openai_api_base=os.environ.get("AZURE_OPENAI_ENDPOINT"),
    #                   openai_api_type="azure",
    #                   temperature= os.environ.get("TEMPERATURE"))
    return llm

# def get_embedding_model():
#     model = OpenAIEmbeddings(deployment=os.environ.get("AZURE_OPENAI_EMBEDDINGS_DEPLOYMENT"),
#                                                openai_api_key=os.environ.get("AZURE_OPENAI_EMBEDDINGS_KEY"),
#                                                openai_api_version=os.environ.get("AZURE_OPENAI_VERSION"),
#                                                openai_api_base=os.environ.get("AZURE_OPENAI_EMBEDDINGS_ENDPOINT"),
#                                                openai_api_type="azure")
#     return model

# def match_title(screen_file,paper):
#     # Read in Screened spreadsheet
#     df = pd.read_excel(screen_file, sheet_name=os.environ.get("SCREEN_NAME") + "_summary")
#     df = df[df['Accept'] == 'Yes']

#     # Assuming your titles are in a column named 'Title'
#     titles = df['Title'].tolist()

#     match_found = False
#     proj_location = os.environ.get("PROJ_LOCATION")
#     pdf_location = proj_location + '/' + os.environ.get('PDF_LOCATION')
#     pdf_title = paper.replace(pdf_location,"").lower().split(" - ")[-1].replace('.pdf','')
#     length = len(pdf_title)
#     for title in titles:
#         score = fuzz.ratio(pdf_title,title.lower()[:length])
        
#         # You can adjust the threshold as needed
#         if score >= 90:  # You can adjust the similarity threshold (e.g., 80)
#             match_found = True
#             print(title)
#             break  # Exit the loop once a match is found
#     return match_found

# def get_abstract(truncated_title, dataframe):
#     # Use fuzzy string matching to find the best match from the "Title" column
#     best_match = process.extractOne(truncated_title, dataframe['Title'])

#     # Check if the fuzzy score exceeds a certain threshold (e.g., 80)
#     if best_match[1] >= 80:
#         # Find the index of the best match
#         best_match_index = dataframe[dataframe['Title'] == best_match[0]].index[0]

#         # Retrieve the corresponding abstract using the index
#         relevant_abstract = dataframe.loc[best_match_index, 'Abstract']
#         return relevant_abstract
#     else:
#         # Handle cases where there's no good match
#         return None

# def colors(integer):
#     color_permutations = [
#         (0.1, 0.5, 0.8),  # Blue shade
#         (0.7, 0.2, 0.1),  # Red shade
#         (0.2, 0.6, 0.3),  # Green shade
#         (0.8, 0.7, 0.1),  # Yellow shade
#         (0.5, 0.2, 0.7),  # Purple shade
#         # Add more color permutations as needed
#     ]
    
#     selected_color = color_permutations[integer % len(color_permutations)]
#     stroke_color = selected_color
#     color_dict = {
#         "stroke": stroke_color,
#     }
#     return color_dict


# def highlight_PDF(paper, phrases, paper_highlight,q_num):
#    # paper_highlight = paper_highlight
#     if os.path.exists(paper_highlight):
#         doc = fitz.open(paper_highlight)
#         incremental = True
#     else:
#         doc = fitz.open(paper)
#         incremental = False
    
#     for page in doc:
#         for phrase in phrases:            
#             text_instances = page.search_for(phrase)

#             for inst in text_instances:
#                 highlight = page.add_highlight_annot(inst)
#                 highlight.set_colors(colors(q_num))
#                 highlight.update()
#     if incremental:
#         doc.save(paper_highlight, incremental=True, encryption=fitz.PDF_ENCRYPT_KEEP)
#     else:
#         doc.save(paper_highlight)

def save_row(identity,df,out_path):
    # Define the file path
    file_name = out_path + '/Extraction_ai.xlsx'
    print("Saving to " + file_name)
    if not os.path.exists(file_name):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = identity
        workbook.save(file_name)
        workbook.close()
    # Write the DataFrame to a new sheet
    workbook = openpyxl.load_workbook(filename=file_name)
    sheet_name = identity
    if sheet_name not in workbook.sheetnames:
        # If the sheet doesn't exist, create it
        sheet = workbook.create_sheet(title=sheet_name)
    else:
        sheet = workbook[sheet_name]
    print("\n" + f"Saving Outputs to Sheet {sheet_name}" + "\n")
    sheet.append(df)
    workbook.save(file_name)
    workbook.close()

# def save_sheet(identity,df,out_path):

#   ### Save Data
#   # Define the file path
#   file_name = out_path + '.xlsx'
#   print("Saving to " + file_name)
#   if not os.path.exists(file_name):
#       workbook = openpyxl.Workbook()
#       workbook.save(file_name)
#       # Open the Excel file using Pandas ExcelWriter
#       excel_writer = pd.ExcelWriter(file_name, engine='openpyxl')

#       # Write the DataFrame to a new sheet
#       sheet_name = identity + '-' + str(1)
#       df.to_excel(excel_writer, sheet_name=sheet_name, index=False)
#       # Save the changes to the Excel file
#       excel_writer._save()
#       print("\n" + f"Saving Outputs to Sheet {sheet_name}" + "\n")
#   else:
#       # If the file exists, open it using openpyxl and add the DataFrame to a new sheet
#       workbook = openpyxl.load_workbook(filename=file_name)
#       sheet_name = identity
#       counter = 0
#       while sheet_name in workbook.sheetnames:
#           counter += 1
#           sheet_name = identity + '-' + str(counter)
#       print("\n" + f"Saving Outputs to Sheet {sheet_name}" + "\n")
#       excel_writer = pd.ExcelWriter(file_name, engine='openpyxl', mode='a')
#       df.to_excel(excel_writer, sheet_name=sheet_name, index=False)
#       excel_writer._save()
#       excel_writer.close()



def generate_text(openAI_key, prompt, n_agents):
    openai.api_key = openAI_key
    #configure_azure_support()
    failed = True
    attempt = 0
    while failed:
        attempt += 1
        if attempt < int(os.environ.get("N_RETRIES")):
            try:
            #  if :#model_to_use == 'gpt-3.5-turbo-instruct':
                # completions = openai.ChatCompletion.create(
                #     engine = os.environ.get("AZURE_OPENAI_DEPLOYMENT"),
                #     #model=model_to_use,
                #     #messages=messages,
                #     max_tokens=512,
                #     prompt = prompt,
                #    # logprobs = 1,
                #     n=n_agents,
                #     stop=None,
                #     temperature= int(os.environ.get("TEMPERATURE")))
                #     # 
            # else:
                messages = [{'role': 'system', 'content': 'You are a helpful assistant.'},
                            {'role': 'user', 'content': prompt}]
                completions = openai.ChatCompletion.create(
                    engine=os.environ.get("AZURE_OPENAI_DEPLOYMENT"),
                    model = os.environ.get("EXTRACTION_MODEL_TO_USE"),
                    messages=messages,
                    max_tokens=512,
                    n=n_agents,
                    stop=None,
                    temperature= int(os.environ.get("TEMPERATURE")))
                failed = False
            except:
                print('Connection Error - Retrying')
                time.sleep(1*2^attempt)     
        else:    
            continue
   # message = completions.choices.message['content']
    return completions
import fitz  # PyMuPDF

class SimpleDocument:
    def __init__(self, page_content):
        self.page_content = page_content
        self.metadata = {'type': None}

def full_text(paper):
    reader = PdfReader(paper)
    raw_documents = [page.extract_text() for page in reader.pages]

    # Wrap raw text in SimpleDocument
    documents = [SimpleDocument(doc) for doc in raw_documents if doc]


    # Process PDF
    text_splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=0)
    clean_docs = text_splitter.split_documents(documents)
    text_chunks = separate_text(clean_docs)
    main_documents = [doc for doc in text_chunks if doc.metadata['type'] == 'main']
    
    all_elements = []
    for document in main_documents:
        all_elements.append(document.page_content)
    
    full_text = ' '.join(all_elements)
    return full_text
from collections import Counter
import re

def separate_text(documents):
    in_frontmatter = True
    in_backmatter = False

    filtered_headers_to_remove = count_line_frequencies(documents)

    for doc in documents:
        new_text = doc.page_content

        # Check and handle backmatter
        if in_backmatter:
            doc.page_content = remove_headers(new_text, filtered_headers_to_remove)
            doc.metadata['type'] = 'backmatter'
            continue

        # Check and handle frontmatter
        if in_frontmatter:
            abstract_position = new_text.lower().find('abstract')
            if abstract_position != -1:
                split_at_abstract = [new_text[:abstract_position], new_text[abstract_position + len('abstract'):]]
                in_frontmatter = False
                doc.page_content = remove_headers(split_at_abstract[0], filtered_headers_to_remove)
                doc.metadata['type'] = 'frontmatter'
                new_text = split_at_abstract[1]
            else:
                split_at_abstract = [new_text]

        # Check and handle main content and references
        split_at_reference = split_at_references(new_text)
        if len(split_at_reference) > 1:
            in_backmatter = True
            new_text = split_at_reference[0]
        
        doc.page_content = remove_headers(new_text, filtered_headers_to_remove)
        doc.metadata['type'] = 'main'

    return documents


def remove_headers(text,filtered_headers_to_remove):

    lines = text.split('\n')

    cleaned_lines = [line for line in lines if all(keyword.lower() not in line.lower() for keyword in filtered_headers_to_remove)]

    return '\n'.join(cleaned_lines)

def count_line_frequencies(documents):

    line_counter = Counter()

    for doc in documents:

        lines = doc.page_content.split('\n')

        line_counter.update(lines)
    
    headers_to_remove = [line for line, count in line_counter.items() if count >= 4]

    filtered_headers_to_remove = [header for header in headers_to_remove if header and not header.isnumeric() and len(header.split()) > 1]

    return filtered_headers_to_remove

def split_at_references(text):
    # Search for various terms
    matches = list(re.finditer(r'(?<=\n)(References|Bibliography|Work Cited|Works Cited|Source List|Literature Cited|Citation List)', text, re.IGNORECASE))
    
    # If found, use the last occurrence for splitting
    if matches:
        last_match = matches[-1]
        before_references = text[:last_match.start()]
        references = text[last_match.start():]
        return before_references, references
    else:
        return [text]
    

def normalize_string(s):
    # Remove punctuation and convert to lowercase
    return re.sub(r'[^a-zA-Z0-9]', '', s).lower().replace("\n", " ")

def strings_similar(str1, str2, threshold=90):
    # Compare two strings with a similarity threshold
    return fuzz.partial_ratio(str1, str2) > threshold



def text_match(context,contents):
    passages = re.split(r'[^\w\s]', context)
    # print(context)
    matches = []
    for passage in passages:
        if len(passage) < 30:
            continue
        # Normalize strings
        normalized_str1 = normalize_string(passage)
        normalized_str2 = normalize_string(contents)

        # Check if one string is in the other
        contained = normalized_str1 in normalized_str2

        # Check for similarity (useful for spelling differences)
        similar = strings_similar(normalized_str1, normalized_str2)

        vectorizer = CountVectorizer()

        # Transform the strings into a document-term matrix
        matrix = vectorizer.fit_transform([normalized_str1, normalized_str2])

        # Calculate the cosine similarity between the two strings
        similarity_score = cosine_similarity(matrix)

        match = contained or similar or similarity_score[0, 1] > 0.5
        matches.append(match)
        print(similarity_score[0, 1])
        if matches == []:
            matches = [True]
    return matches

def read_first_entries(directory,num_entries):
    file_data = []
    for filename in os.listdir(directory):
        if filename.endswith('.csv'):
            # Read the CSV file
            df = pd.read_csv(os.path.join(directory, filename))
            
            # Check if the first column has at least 14 elements
            if len(df) >= num_entries:
                # Store the first 14 elements of the first column
                elements = df.iloc[:num_entries, 0].tolist()
            else:
                # Store all elements if less than 14
                elements = df.iloc[:, 0].tolist()
        file_data.extend(elements)
    return file_data

# Replace 'your_directory_path' with the path to your directory containing CSV files

def ask_AI(prompt,AI_to_use):
    n_attempts = 0
    
    if 'azure' in AI_to_use:
        client = AzureChatOpenAI(azure_endpoint=os.environ.get("AZURE_OPENAI_ENDPOINT"),
                     api_key=os.environ.get("AZURE_OPENAI_KEY"),
                     api_version=os.environ.get("AZURE_OPENAI_VERSION"),
                     azure_deployment=os.environ.get("AZURE_OPENAI_DEPLOYMENT"))
    else:
        client = ChatOpenAI(api_key=os.environ.get("OPENAI_API_KEY"))
    if 'gpt-4-turbo' in AI_to_use:
        model = 'gpt-4-0125-preview'
    elif 'mixtral' in AI_to_use:
        model = "mixtral-8x7b-32768"
    elif 'llama' in AI_to_use:
        model = 'llama2-70b-4096'
    else:
        model = 'gpt-3.5-turbo-0125'
    while n_attempts < 5:
        try:
            completion = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": "You are a helpful assistant."},
                    {"role": "user", "content": prompt}
                ]
                )
            n_attempts = 999
        except:
            n_attempts += 1
    output = completion.choices[0].message.content
    return(output)